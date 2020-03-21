from datetime import datetime

import openpyxl
from openpyxl.utils import get_column_letter


class ExcelWriter:
    def __init__(self, filename):
        self.filename = filename
        self._wb = openpyxl.Workbook()

    def create_workbook(self):
        self._wb = openpyxl.Workbook()

    def create_worksheet(self, name):
        self._wb.create_sheet(name)

    @property
    def workbook(self):
        return self._wb

    @property
    def worksheet(self):
        return self._wb.active

    def write_row(self, row, column, row_data):
        for shift, value in enumerate(row_data):
            self.worksheet.cell(row, column + shift, value)

    def write_rows(self, row, column, rows_data):
        for shift, value in enumerate(rows_data):
            self.write_row(row + shift, column, value)

    def save(self):
        self.workbook.save(self.filename)

    def close(self):
        self.workbook.close()

    def save_and_close(self):
        self.save()
        self.close()


def apply_sheet_format(sheet):
    max_row = sheet.max_row
    max_column = sheet.max_column

    exchange_rate_col = []
    for col in range(1, max_column + 1):
        column = [sheet.cell(row=row, column=col) for row in range(1, max_row + 1)]

        if column[0].value == 'Дата':
            # Формат чисел – дата
            for cell_data in column[1:]:
                cell_data.number_format = 'dd/mm/yy;@'

        if column[0].value == 'Курс':
            # Формат чисел – финансовый
            for cell_data in column[1:]:
                cell_data.number_format = '_-* #,##0.00\ "₽"_-;\-* #,##0.00\ "₽"_-;_-* "-"??\ "₽"_-;_-@_-'
            # Добавить номера ячеек с курсами USD, EUR
            exchange_rate_col.append([rate_cell.coordinate for rate_cell in column[1:]])

        if column[0].value == 'Изменение':
            # Формат чисел – финансовый
            for cell_data in column[1:]:
                cell_data.number_format = '_-* #,##0.00\ "₽"_-;\-* #,##0.00\ "₽"_-;_-* "-"??\ "₽"_-;_-@_-'

    # Ячейки в ближайшем свободном столбце
    ratio = [sheet.cell(row=row, column=max_column + 1) for row in range(2, max_row + 1)]

    # Добавить формулу: поделить курс евро на доллар
    for cell, cell_ratio in enumerate(ratio, 2):
        cell_ratio.value = f'={exchange_rate_col[1][cell - 2]}/{exchange_rate_col[0][cell - 2]}'

    # Автоширина
    for col in range(1, sheet.max_column + 1):
        sheet.column_dimensions[get_column_letter(col)].bestFit = True

    return sheet
