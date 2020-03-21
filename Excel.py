import openpyxl
from openpyxl.utils import get_column_letter

from Action import Action


class ExcelWriter:
    def __init__(self, filename):
        self.filename = filename
        self._wb = openpyxl.Workbook()

    def create_workbook(self):
        self._wb = openpyxl.Workbook()

    def create_worksheet(self, name):
        self._wb.create_sheet(name)

    @property
    def workbook(self):
        return self._wb

    @property
    def worksheet(self):
        return self._wb.active

    def write_row(self, row, column, row_data):
        for shift, value in enumerate(row_data):
            self.worksheet.cell(row, column + shift, value)

    def write_rows(self, row, column, rows_data):
        for shift, value in enumerate(rows_data):
            self.write_row(row + shift, column, value)

    @property
    def next_free_column(self):
        max_col = self.worksheet.max_column
        return max_col if max_col == 1 else max_col + 1

    def save(self):
        self.workbook.save(self.filename)

    def close(self):
        self.workbook.close()

    def save_and_close(self):
        self.save()
        self.close()


class WriteToXlsAction(Action):
    def __init__(self, filename):
        self.writer = ExcelWriter(filename)

    def process(self, context):
        self.write_rows(context.headers, context.rows)
        self.apply_sheet_format(self.writer.worksheet)
        context.rows_count = self.writer.worksheet.max_row
        self.writer.save_and_close()

    def write_rows(self, headers, rows):
        next_free_column = self.writer.next_free_column
        self.writer.write_row(row=1, column=next_free_column, row_data=headers)
        self.writer.write_rows(row=2, column=next_free_column, rows_data=rows)

    def apply_sheet_format(self, sheet):
        max_column = sheet.max_column

        styles = {'Дата': 'dd/mm/yy;@',
                  'Курс': '_-* #,##0.00\ "₽"_-;\-* #,##0.00\ "₽"_-;_-* "-"??\ "₽"_-;_-@_-',
                  'Изменение': '_-* #,##0.00\ "₽"_-;\-* #,##0.00\ "₽"_-;_-* "-"??\ "₽"_-;_-@_-',
                  }

        exchange_rate_col = []
        for column in self.writer.worksheet.iter_cols(min_col=1, max_col=max_column):
            style = styles.get(column[0].value)
            if style is not None:
                self.apply_column_style(column[1:], style)
            if column[0].value == 'Курс':
                exchange_rate_col.append([rate_cell.coordinate for rate_cell in column[1:]])

        # self.write_ratio(exchange_rate_col, max_column, max_row, sheet)

        # Автоширина
        for col in range(1, self.writer.next_free_column):
            sheet.column_dimensions[get_column_letter(col)].bestFit = True

        return sheet

    def apply_column_style(self, column, style):
        for cell in column:
            cell.number_format = style

    def write_ratio(self, exchange_rate_col):
        max_row = self.writer.worksheet.max_row
        sheet = self.writer.worksheet
        # Ячейки в ближайшем свободном столбце

        ratio = [sheet.cell(row=row, column=self.writer.next_free_column) for row in range(2, max_row + 1)]
        # Добавить формулу: поделить курс евро на доллар
        for cell, cell_ratio in enumerate(ratio, 2):
            cell_ratio.value = f'={exchange_rate_col[1][cell - 2]}/{exchange_rate_col[0][cell - 2]}'
