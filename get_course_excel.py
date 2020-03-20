# -*- coding: utf-8 -*-
import requests
from bs4 import BeautifulSoup
from time import sleep
import datetime
from os import path
import openpyxl
from openpyxl.utils import get_column_letter


# Получаем ссылку на динамику курса
def get_link_currency(url, currency):
    soup = get_page_soup(url)
    link = soup.find(text=currency).parent.attrs['href']
    return link


# Получаем данные динамики курса
def get_quote_data(url):
    table = []
    soup = get_page_soup(url)

    table.append([cell.get_text(strip=True) for cell in soup.select('th')])

    for row in soup.select('tr'):
        cells = row.find_all('td')
        if not cells:
            continue
        table.append([cell.get_text(strip=True) for cell in cells])

    return table


# Сохранение полученных данных в xls
def save_to_excel(name_file, data):
    if not path.isfile(name_file):
        wb = openpyxl.Workbook()
        wb.create_sheet(title='Тестовое задание', index=0)
    else:
        wb = openpyxl.load_workbook(name_file)
    sheet = wb['Тестовое задание']

    max_column = sheet.max_column
    if max_column > 1:
        max_column += 1

    headers = data[0]
    values = data[1:]
    # Записать заголовки
    for n_col, header in enumerate(headers, 0):
        cell_headers = sheet.cell(row=1, column=max_column + n_col)
        cell_headers.value = header
    # Записать данные по столбцам
    for n_row, row in enumerate(values,2):
        cell_1 = sheet.cell(row=n_row, column=max_column)
        cell_1.value = datetime.datetime.strptime(f"{row[0]}", "%d.%m.%y").date()
        cell_2 = sheet.cell(row=n_row, column=max_column + 1)
        cell_2.value = float(row[1].replace(',', '.'))
        cell_3 = sheet.cell(row=n_row, column=max_column + 2)
        cell_3.value = float(row[2].replace(',', '.'))

    wb.save(name_file)
    wb.close()


# Форматирование сохранёных данных в xls
def format_excel(name_file):
    wb = openpyxl.load_workbook(name_file)
    wb.active = 0

    sheet = wb.active

    max_row = sheet.max_row
    max_column = sheet.max_column

    exchange_rate_col = []
    for col in range(1, max_column+1):
        column = [sheet.cell(row=row, column=col) for row in range(1, max_row + 1)]

        if column[0].value == 'Дата':
            # Формат чисел – дата
            for cell_data in column[1:]:
                cell_data.number_format = 'dd/mm/yy;@'

        if column[0].value == 'Курс':
            # Формат чисел – финансовый
            for cell_data in column[1:]:
                cell_data.number_format = '_-* #,##0.00\ "₽"_-;\-* #,##0.00\ "₽"_-;_-* "-"??\ "₽"_-;_-@_-'
            # Добавить номера ячеек с курсами USD, EUR
            exchange_rate_col.append([rate_cell.coordinate for rate_cell in column[1:]])

        if column[0].value == 'Изменение':
            # Формат чисел – финансовый
            for cell_data in column[1:]:
                cell_data.number_format = '_-* #,##0.00\ "₽"_-;\-* #,##0.00\ "₽"_-;_-* "-"??\ "₽"_-;_-@_-'

    # Ячейки в ближайшем свободном столбце
    ratio = [sheet.cell(row=row, column=max_column+1) for row in range(2, max_row+1)]

    # Добавить формулу: поделить курс евро на доллар
    for cell, cell_ratio in enumerate(ratio, 2):
        cell_ratio.value = f'={exchange_rate_col[1][cell-2]}/{exchange_rate_col[0][cell-2]}'

    # Автоширина
    for col in range(1, sheet.max_column + 1):
        sheet.column_dimensions[get_column_letter(col)].bestFit = True

    wb.save(name_file)
    wb.close()
    return max_row


# Получаем страницу
def get_page_soup(url):
    headers_response = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.132 Safari/537.36'}
    response = requests.get(url, headers=headers_response)
    soup = BeautifulSoup(response.text, features="lxml")
    sleep(1)
    return soup




